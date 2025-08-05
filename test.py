
try:
    import camelot
    CAMELOT_AVAILABLE = True
    print("✓ camelot-py 사용 가능합니다.")
except ImportError:
    CAMELOT_AVAILABLE = False
    print("⚠ camelot-py를 사용할 수 없습니다. pdfplumber만 사용합니다.")

import pdfplumber
import pandas as pd
import re
import html
import requests
import time
import subprocess
import os
import tempfile
import datetime
from openpyxl import load_workbook

# DeepL API 설정
DEEPL_API_KEY = os.getenv("DEEPL_API_KEY")
if not DEEPL_API_KEY:
    raise ValueError("DEEPL_API_KEY 환경 변수가 설정되지 않았습니다.")
DEEPL_API_URL = "https://api-free.deepl.com/v2/translate"
OCR_LANGUAGE = "eng+kor"

def print_progress(message, is_header=False):
    """진행 상황 출력"""
    timestamp = datetime.datetime.now().strftime("%H:%M:%S")
    if is_header:
        print("=" * 60)
        print(f"[{timestamp}] {message}")
        print("=" * 60)
    else:
        print(f"[{timestamp}] {message}")

def get_unique_filename(directory, base_name, extension=".xlsx"):
    """고유한 파일명 생성"""
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]
    filename = f"{base_name}_{timestamp}_final{extension}"
    return os.path.join(directory, filename)

def find_pdf_files(directory):
    """디렉토리에서 모든 PDF 파일 찾기"""
    pdf_files = []
    for filename in os.listdir(directory):
        if filename.lower().endswith('.pdf'):
            pdf_files.append(os.path.join(directory, filename))
    return sorted(pdf_files)

def check_ocr_dependencies():
    """OCR 의존성 확인 - 핵심 의존성만 체크"""
    print_progress("OCR 의존성 확인 중...")
    
    # 핵심 의존성만 체크 (ghostscript는 선택사항)
    core_dependencies = {
        'ocrmypdf': ['ocrmypdf', '--version'],
        'tesseract': ['tesseract', '--version']
    }
    
    optional_dependencies = {
        'ghostscript': ['gs', '--version']
    }
    
    missing_core_deps = []
    
    # 핵심 의존성 체크
    for name, cmd in core_dependencies.items():
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
            if result.returncode == 0:
                version = result.stdout.strip().split('\n')[0]
                print_progress(f"✓ {name}: {version}")
            else:
                print_progress(f"✗ {name}: 설치되어 있지만 실행 실패")
                missing_core_deps.append(name)
        except FileNotFoundError:
            print_progress(f"✗ {name}: 설치되지 않음")
            missing_core_deps.append(name)
        except Exception as e:
            print_progress(f"✗ {name}: 확인 중 오류 - {str(e)}")
            missing_core_deps.append(name)
    
    # 선택적 의존성 체크 (실패해도 계속 진행)
    for name, cmd in optional_dependencies.items():
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
            if result.returncode == 0:
                version = result.stdout.strip().split('\n')[0]
                print_progress(f"✓ {name}: {version}")
            else:
                print_progress(f"⚠ {name}: 설치되어 있지만 실행 실패 (OCR 성능에 영향 있을 수 있음)")
        except FileNotFoundError:
            print_progress(f"⚠ {name}: 설치되지 않음 (OCR 성능에 영향 있을 수 있음)")
        except Exception as e:
            print_progress(f"⚠ {name}: 확인 중 오류 - {str(e)} (OCR 성능에 영향 있을 수 있음)")
    
    if missing_core_deps:
        print_progress("누락된 핵심 의존성 설치 방법 (Windows):")
        for dep in missing_core_deps:
            if dep == 'ocrmypdf':
                print_progress("  pip install ocrmypdf")
            elif dep == 'tesseract':
                print_progress("  Windows: https://github.com/UB-Mannheim/tesseract/wiki")
                print_progress("  또는 Chocolatey: choco install tesseract")
        return False
    else:
        print_progress("✓ 핵심 OCR 의존성이 설치되어 있습니다.")
        return True

def check_tesseract_languages():
    """Tesseract 언어팩 확인"""
    try:
        result = subprocess.run(['tesseract', '--list-langs'], 
                               capture_output=True, text=True, timeout=10)
        if result.returncode == 0:
            available_langs = result.stdout.strip().split('\n')[1:]  # 첫 줄은 헤더
            print_progress(f"사용 가능한 언어: {', '.join(available_langs[:10])}")  # 처음 10개만 표시
            
            # 한국어와 영어 확인
            has_korean = 'kor' in available_langs
            has_english = 'eng' in available_langs
            
            if has_korean and has_english:
                print_progress("✓ 한국어, 영어 언어팩 모두 사용 가능")
                return 'eng+kor'
            elif has_english:
                print_progress("⚠ 영어 언어팩만 사용 가능 (한국어 없음)")
                return 'eng'
            else:
                print_progress("✗ 영어 언어팩도 없음")
                return None
        else:
            print_progress("✗ Tesseract 언어 목록 확인 실패")
            return 'eng'  # 기본값
    except Exception as e:
        print_progress(f"✗ 언어팩 확인 중 오류: {str(e)}")
        return 'eng'  # 기본값

def check_pdf_has_tables_and_text(input_pdf):
    """PDF에 테이블과 텍스트가 있는지 사전 확인"""
    has_tables = False
    has_text = False
    total_text_length = 0
    
    # 테이블 확인 - camelot이 사용 가능한 경우에만
    if CAMELOT_AVAILABLE:
        try:
            tables = camelot.read_pdf(input_pdf, pages='all')
            if tables and len(tables) > 0:
                valid_tables = 0
                for table in tables:
                    if table.df.shape[0] > 1 and table.df.shape[1] > 1:
                        non_empty_cells = table.df.count().sum()
                        total_cells = table.df.shape[0] * table.df.shape[1]
                        if non_empty_cells / total_cells > 0.1:
                            valid_tables += 1
                
                if valid_tables > 0:
                    has_tables = True
                    print_progress(f"✓ 유효한 테이블 {valid_tables}개 발견 (camelot)")
        except Exception as e:
            print_progress(f"⚠ camelot 테이블 추출 실패: {str(e)}")
    
    # pdfplumber로 테이블 감지 시도
    if not has_tables:
        try:
            with pdfplumber.open(input_pdf) as pdf:
                table_count = 0
                for page in pdf.pages:
                    page_tables = page.find_tables()
                    if page_tables and len(page_tables) > 0:
                        table_count += len(page_tables)
                
                if table_count > 0:
                    has_tables = True
                    print_progress(f"✓ {table_count}개 테이블 발견 (pdfplumber)")
        except Exception as e:
            print_progress(f"⚠ pdfplumber 테이블 감지 실패: {str(e)}")
    
    # 텍스트 확인 - 더 포괄적인 방법 사용
    try:
        with pdfplumber.open(input_pdf) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                text_methods = [
                    page.extract_text(),
                    page.extract_text(layout=True),
                    page.extract_text(x_tolerance=3, y_tolerance=3),
                    page.extract_text(x_tolerance=1, y_tolerance=1),
                    page.extract_text(x_tolerance=0, y_tolerance=0)
                ]
                
                # 문자 단위로 직접 추출도 시도
                try:
                    chars = page.chars
                    if chars:
                        char_text = ''.join([char['text'] for char in chars])
                        text_methods.append(char_text)
                except:
                    pass
                
                # 단어 단위로 추출도 시도
                try:
                    words = page.extract_words()
                    if words:
                        word_text = ' '.join([word['text'] for word in words])
                        text_methods.append(word_text)
                except:
                    pass
                
                page_text_found = False
                for text in text_methods:
                    if text:
                        clean_text = text.strip()
                        unique_chars = len(set(clean_text.replace(' ', '').replace('\n', '')))
                        if len(clean_text) > 20 and unique_chars > 5:  # 기준을 낮춤
                            total_text_length += len(clean_text)
                            has_text = True
                            page_text_found = True
                            print_progress(f"페이지 {page_num}: {len(clean_text)} 문자 발견")
                            break
                
                if not page_text_found:
                    print_progress(f"페이지 {page_num}: 텍스트 없음 (이미지 페이지일 가능성)")
        
        if has_text:
            print_progress(f"✓ 텍스트 발견 (총 {total_text_length} 문자)")
        else:
            print_progress("✗ 의미있는 텍스트 없음 - 이미지 기반 PDF일 가능성")
            
    except Exception as e:
        print_progress(f"✗ 텍스트 추출 실패: {str(e)}")
    
    # 종합 점수 계산
    content_score = 0
    if has_tables:
        content_score += 100
    if has_text and total_text_length > 100:
        content_score += 50
    elif has_text:
        content_score += 20
    
    return has_tables, has_text, total_text_length, content_score

def ocr_pdf_with_options(input_pdf, language='eng+kor'):
    """개선된 OCR 처리 - 상세 디버깅 포함"""
    try:
        # 1. OCR 의존성 확인 (관대한 모드)
        deps_ok = check_ocr_dependencies()
        if not deps_ok:
            print_progress("⚠ 일부 OCR 의존성이 부족하지만 OCR을 시도해보겠습니다.")
        else:
            print_progress("✓ OCR 의존성 확인 완료")
        
        # 2. 언어팩 확인 및 조정
        available_language = check_tesseract_languages()
        if not available_language:
            print_progress("✗ 사용 가능한 언어팩이 없습니다. OCR을 건너뜁니다.")
            return input_pdf
        
        if available_language != language:
            print_progress(f"언어 설정 변경: {language} → {available_language}")
            language = available_language
        
        # 3. 파일 상태 확인
        if not os.path.exists(input_pdf):
            print_progress(f"✗ 입력 파일을 찾을 수 없습니다: {input_pdf}")
            return input_pdf
        
        file_size = os.path.getsize(input_pdf) / (1024 * 1024)  # MB
        print_progress(f"파일 크기: {file_size:.2f} MB")
        
        if file_size > 100:
            print_progress("⚠ 파일이 큰 편입니다. 처리 시간이 오래 걸릴 수 있습니다.")
        
        # 4. 출력 파일명 생성
        base_name = os.path.splitext(input_pdf)[0]
        output_pdf = f"{base_name}_ocr_{int(time.time())}.pdf"
        
        # 5. 고급 OCR 시도
        success = ocr_pdf_advanced(input_pdf, output_pdf, language)
        if success:
            return output_pdf
        
        # 6. 기본 OCR 시도
        print_progress("고급 OCR 실패, 기본 OCR 시도...")
        success = ocr_pdf_basic(input_pdf, output_pdf, language)
        if success:
            return output_pdf
        
        # 7. 최소 OCR 시도
        print_progress("기본 OCR 실패, 최소 OCR 시도...")
        success = ocr_pdf_minimal(input_pdf, output_pdf, language)
        if success:
            return output_pdf
        
        # 8. 가장 간단한 OCR 시도
        print_progress("최소 OCR 실패, 가장 간단한 OCR 시도...")
        success = ocr_pdf_simple(input_pdf, output_pdf, language)
        if success:
            return output_pdf
        
        print_progress("✗ 모든 OCR 시도 실패 - 원본 파일 사용")
        return input_pdf
        
    except Exception as e:
        print_progress(f"✗ OCR 처리 중 예외 발생: {str(e)}")
        return input_pdf

def ocr_pdf_advanced(input_pdf, output_pdf, language):
    """고급 옵션을 사용한 OCR 처리"""
    try:
        cmd = [
            'ocrmypdf',
            '--force-ocr',
            '--optimize', '1',
            '--language', language,
            '--deskew',
            '--clean',
            '--rotate-pages',
            '--remove-background',
            '--oversample', '300',
            '--output-type', 'pdf',
            input_pdf,
            output_pdf
        ]
        
        print_progress("🔍 고급 OCR 처리 중...")
        print_progress(f"명령어: {' '.join(cmd)}")
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=600)
        
        if result.returncode == 0:
            print_progress("✓ 고급 OCR 완료")
            return True
        else:
            print_progress(f"✗ 고급 OCR 실패 (코드: {result.returncode})")
            if result.stderr:
                print_progress(f"오류 메시지: {result.stderr}")
            return False
            
    except subprocess.TimeoutExpired:
        print_progress("✗ 고급 OCR 시간 초과 (10분)")
        return False
    except Exception as e:
        print_progress(f"✗ 고급 OCR 중 오류: {str(e)}")
        return False

def ocr_pdf_basic(input_pdf, output_pdf, language):
    """기본 옵션을 사용한 OCR 처리"""
    try:
        cmd = [
            'ocrmypdf',
            '--force-ocr',
            '--language', language,
            '--optimize', '1',
            input_pdf,
            output_pdf
        ]
        
        print_progress("🔍 기본 OCR 처리 중...")
        print_progress(f"명령어: {' '.join(cmd)}")
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
        
        if result.returncode == 0:
            print_progress("✓ 기본 OCR 완료")
            return True
        else:
            print_progress(f"✗ 기본 OCR 실패 (코드: {result.returncode})")
            if result.stderr:
                print_progress(f"오류 메시지: {result.stderr}")
            return False
            
    except subprocess.TimeoutExpired:
        print_progress("✗ 기본 OCR 시간 초과 (5분)")
        return False
    except Exception as e:
        print_progress(f"✗ 기본 OCR 중 오류: {str(e)}")
        return False

def ocr_pdf_minimal(input_pdf, output_pdf, language):
    """최소 옵션을 사용한 OCR 처리 - Ghostscript 없이도 동작"""
    try:
        cmd = [
            'ocrmypdf',
            '--language', language,
            '--skip-text',  # 기존 텍스트 건너뛰기
            input_pdf,
            output_pdf
        ]
        
        print_progress("🔍 최소 OCR 처리 중...")
        print_progress(f"명령어: {' '.join(cmd)}")
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
        
        if result.returncode == 0:
            print_progress("✓ 최소 OCR 완료")
            return True
        else:
            print_progress(f"✗ 최소 OCR 실패 (코드: {result.returncode})")
            if result.stderr:
                print_progress(f"오류 메시지: {result.stderr}")
            
            # 더 간단한 OCR 시도
            return ocr_pdf_simple(input_pdf, output_pdf, language)
            
    except subprocess.TimeoutExpired:
        print_progress("✗ 최소 OCR 시간 초과 (3분)")
        return False
    except Exception as e:
        print_progress(f"✗ 최소 OCR 중 오류: {str(e)}")
        return False

def ocr_pdf_simple(input_pdf, output_pdf, language):
    """가장 간단한 OCR 처리"""
    try:
        cmd = [
            'ocrmypdf',
            '--force-ocr',  # 강제 OCR
            '--language', language,
            input_pdf,
            output_pdf
        ]
        
        print_progress("🔍 간단 OCR 처리 중...")
        print_progress(f"명령어: {' '.join(cmd)}")
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
        
        if result.returncode == 0:
            print_progress("✓ 간단 OCR 완료")
            return True
        else:
            print_progress(f"✗ 간단 OCR 실패 (코드: {result.returncode})")
            if result.stderr:
                print_progress(f"오류 메시지: {result.stderr}")
            return False
            
    except subprocess.TimeoutExpired:
        print_progress("✗ 간단 OCR 시간 초과 (5분)")
        return False
    except Exception as e:
        print_progress(f"✗ 간단 OCR 중 오류: {str(e)}")
        return False

def translate_with_deepl(text, source_lang="EN", target_lang="KO"):
    """DeepL API를 사용하여 텍스트 번역"""
    if not text or not text.strip():
        return ""
    
    headers = {
        "Authorization": f"DeepL-Auth-Key {DEEPL_API_KEY}",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    
    data = {
        "text": text,
        "source_lang": source_lang,
        "target_lang": target_lang
    }
    
    try:
        response = requests.post(DEEPL_API_URL, headers=headers, data=data)
        
        if response.status_code == 200:
            result = response.json()
            return result["translations"][0]["text"]
        else:
            return text
    except Exception:
        return text

def clean_text_for_excel(text):
    """Excel XML 호환을 위한 텍스트 정리"""
    if not text:
        return ""
    
    text = str(text)
    text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', text)
    text = html.unescape(text)
    text = text.replace('&', '&amp;')
    text = text.replace('<', '&lt;')
    text = text.replace('>', '&gt;')
    text = text.replace('"', '&quot;')
    text = text.replace("'", '&apos;')
    text = re.sub(r'\s+', ' ', text)
    text = text.strip()
    
    return text

def extract_text_improved(text):
    """텍스트 추출 및 정리 함수 개선"""
    if not text:
        return ""
    
    text = text.strip()
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'\n\s*\n', '\n', text)
    cleaned_text = re.sub(r'[^\w\s.,!?;:\-\(\)\[\]\/\\\'"@#$%&+=<>]', ' ', text)
    cleaned_text = re.sub(r'\s+', ' ', cleaned_text)
    cleaned_text = clean_text_for_excel(cleaned_text)
    
    return cleaned_text.strip()

def clean_dataframe(df):
    """DataFrame의 모든 셀을 Excel 호환 형태로 정리"""
    if df.empty:
        return df
    
    cleaned_df = df.copy()
    for col in cleaned_df.columns:
        cleaned_df[col] = cleaned_df[col].apply(
            lambda x: clean_text_for_excel(str(x)) if pd.notna(x) else ""
        )
    
    return cleaned_df

def pdf_to_excel(input_pdf, output_xlsx):
    """PDF에서 텍스트와 테이블을 추출하여 Excel로 변환"""
    try:
        print_progress("📊 테이블 추출 중...")
        # 테이블 추출
        tables = []
        
        # camelot이 사용 가능한 경우 먼저 시도
        if CAMELOT_AVAILABLE:
            try:
                tables = camelot.read_pdf(input_pdf, pages='all')
                if tables:
                    print_progress(f"✓ camelot으로 {len(tables)}개 테이블 추출됨")
            except Exception as e:
                print_progress(f"⚠ camelot 테이블 추출 실패: {str(e)}")
        
        # pdfplumber로 테이블 추출 (camelot 실패 시 또는 camelot 미사용 시)
        pdfplumber_tables = []
        if not tables:
            print_progress("pdfplumber를 사용하여 테이블 추출 시작...")
            try:
                with pdfplumber.open(input_pdf) as pdf:
                    print_progress(f"PDF 총 페이지 수: {len(pdf.pages)}")
                    for page_num, page in enumerate(pdf.pages, 1):
                        print_progress(f"페이지 {page_num} 분석 중... (크기: {page.width:.1f} x {page.height:.1f})")
                        
                        # 페이지에 라인이나 도형이 있는지 확인
                        lines = page.lines if hasattr(page, 'lines') else []
                        rects = page.rects if hasattr(page, 'rects') else []
                        curves = page.curves if hasattr(page, 'curves') else []
                        
                        print_progress(f"페이지 {page_num}: 라인 {len(lines)}개, 사각형 {len(rects)}개, 곡선 {len(curves)}개")
                        
                        # 텍스트 밀도 확인
                        words = page.extract_words()
                        if words:
                            print_progress(f"페이지 {page_num}: 단어 {len(words)}개 발견")
                        try:
                            # 여러 방법으로 테이블 찾기 시도
                            page_tables = page.find_tables()
                            
                            # 기본 설정으로 테이블을 찾지 못한 경우, 더 관대한 설정들을 시도
                            if not page_tables:
                                table_settings_list = [
                                    # 설정 1: 라인 기반, 엄격
                                    {
                                        "vertical_strategy": "lines_strict",
                                        "horizontal_strategy": "lines_strict",
                                        "min_words_vertical": 1,
                                        "min_words_horizontal": 1
                                    },
                                    # 설정 2: 라인 기반, 관대
                                    {
                                        "vertical_strategy": "lines",
                                        "horizontal_strategy": "lines",
                                        "min_words_vertical": 1,
                                        "min_words_horizontal": 1
                                    },
                                    # 설정 3: 텍스트 기반
                                    {
                                        "vertical_strategy": "text",
                                        "horizontal_strategy": "text",
                                        "min_words_vertical": 2,
                                        "min_words_horizontal": 2,
                                        "intersection_tolerance": 5
                                    },
                                    # 설정 4: 매우 관대한 설정
                                    {
                                        "vertical_strategy": "explicit",
                                        "horizontal_strategy": "explicit",
                                        "explicit_vertical_lines": [],
                                        "explicit_horizontal_lines": [],
                                        "min_words_vertical": 1,
                                        "min_words_horizontal": 1,
                                        "intersection_tolerance": 10
                                    }
                                ]
                                
                                for i, settings in enumerate(table_settings_list):
                                    try:
                                        page_tables = page.find_tables(table_settings=settings)
                                        if page_tables:
                                            print_progress(f"페이지 {page_num}: 설정 {i+1}로 {len(page_tables)}개 테이블 발견")
                                            break
                                    except Exception as e:
                                        print_progress(f"페이지 {page_num}: 설정 {i+1} 실패 - {str(e)}")
                                        continue
                            
                            if page_tables:
                                for table_idx, table in enumerate(page_tables):
                                    try:
                                        # 더 안전한 테이블 추출
                                        table_data = None
                                        
                                        # 방법 1: 기본 extract_table
                                        try:
                                            table_data = page.extract_table(table.bbox)
                                        except:
                                            pass
                                        
                                        # 방법 2: 수동으로 bbox 조정
                                        if not table_data:
                                            try:
                                                bbox = table.bbox
                                                # bbox 좌표를 약간 조정
                                                adjusted_bbox = (
                                                    max(0, bbox[0] - 5),
                                                    max(0, bbox[1] - 5), 
                                                    min(page.width, bbox[2] + 5),
                                                    min(page.height, bbox[3] + 5)
                                                )
                                                table_data = page.extract_table(adjusted_bbox)
                                            except:
                                                pass
                                        
                                        # 방법 3: within_bbox 사용
                                        if not table_data:
                                            try:
                                                cropped_page = page.within_bbox(table.bbox)
                                                table_data = cropped_page.extract_table()
                                            except:
                                                pass
                                        
                                        if table_data and len(table_data) > 1:
                                            # 빈 행 제거 및 정리
                                            filtered_data = []
                                            for row in table_data:
                                                if row and any(cell and str(cell).strip() for cell in row):
                                                    # None 값을 빈 문자열로 변경
                                                    clean_row = [str(cell).strip() if cell else "" for cell in row]
                                                    filtered_data.append(clean_row)
                                            
                                            if len(filtered_data) > 0:
                                                # 헤더 설정
                                                if len(filtered_data) > 1:
                                                    headers = filtered_data[0] if any(filtered_data[0]) else [f"Col_{i}" for i in range(len(filtered_data[0]))]
                                                    data_rows = filtered_data[1:]
                                                else:
                                                    headers = [f"Col_{i}" for i in range(len(filtered_data[0]))]
                                                    data_rows = filtered_data
                                                
                                                # 모든 행의 길이를 헤더 길이에 맞춤
                                                for i, row in enumerate(data_rows):
                                                    while len(row) < len(headers):
                                                        row.append("")
                                                    data_rows[i] = row[:len(headers)]
                                                
                                                df = pd.DataFrame(data_rows, columns=headers)
                                                
                                                # camelot 형식과 호환되도록 객체 생성
                                                class PDFPlumberTable:
                                                    def __init__(self, df, page_num):
                                                        self.df = df
                                                        self.page = page_num
                                                
                                                pdfplumber_tables.append(PDFPlumberTable(df, page_num))
                                                print_progress(f"✓ 페이지 {page_num}에서 테이블 {table_idx + 1} 추출됨 ({df.shape[0]}행 {df.shape[1]}열)")
                                            else:
                                                print_progress(f"⚠ 페이지 {page_num} 테이블 {table_idx + 1}: 유효한 데이터 없음")
                                        else:
                                            print_progress(f"⚠ 페이지 {page_num} 테이블 {table_idx + 1}: 추출된 데이터 없음")
                                    
                                    except Exception as table_error:
                                        print_progress(f"⚠ 페이지 {page_num} 테이블 {table_idx + 1} 추출 실패: {str(table_error)}")
                                        continue
                            else:
                                print_progress(f"페이지 {page_num}: 테이블 없음")
                        
                        except Exception as page_error:
                            print_progress(f"⚠ 페이지 {page_num} 처리 실패: {str(page_error)}")
                            continue
                
                if pdfplumber_tables:
                    tables = pdfplumber_tables
                    print_progress(f"✓ pdfplumber로 총 {len(tables)}개 테이블 추출됨")
            except Exception as e:
                print_progress(f"⚠ pdfplumber 테이블 추출 실패: {str(e)}")
        
        if not tables:
            print_progress("⚠ 테이블을 찾을 수 없습니다. 텍스트만 추출합니다.")
        
        print_progress("📄 텍스트 추출 중...")
        # 텍스트 추출
        page_texts = {}
        total_text_length = 0
        
        with pdfplumber.open(input_pdf) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                text_methods = [
                    page.extract_text(),
                    page.extract_text(layout=True),
                    page.extract_text(x_tolerance=3, y_tolerance=3),
                    page.extract_text(y_tolerance=3)
                ]
                
                best_text = ""
                for text in text_methods:
                    if text and len(text) > len(best_text):
                        best_text = text
                
                if best_text:
                    cleaned_text = extract_text_improved(best_text)
                    page_texts[page_num] = cleaned_text
                    total_text_length += len(cleaned_text)
                    print_progress(f"페이지 {page_num}: {len(cleaned_text)} 문자 추출됨")
                else:
                    print_progress(f"페이지 {page_num}: 텍스트 추출 불가")
        
        print_progress(f"✓ {len(page_texts)}개 페이지에서 {total_text_length} 문자 추출됨")
        
        # 페이지별로 테이블 그룹화
        page_tables = {}
        if tables:
            for table in tables:
                page_num = table.page
                if page_num not in page_tables:
                    page_tables[page_num] = []
                page_tables[page_num].append(table)
        
        print_progress("📋 Excel 파일 생성 중...")
        # Excel 파일 생성
        with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
            all_pages = set(page_tables.keys()) | set(page_texts.keys())
            
            if not all_pages:
                all_pages = {1}
                page_texts[1] = "PDF에서 내용을 추출할 수 없습니다."
            
            for page_num in sorted(all_pages):
                sheet_data = []
                
                # 페이지 텍스트 추가
                if page_num in page_texts and page_texts[page_num]:
                    text_content = page_texts[page_num]
                    text_lines = text_content.split('\n')
                    
                    for line in text_lines:
                        line = line.strip()
                        if line:
                            cleaned_line = clean_text_for_excel(line)
                            if cleaned_line:
                                sheet_data.append([cleaned_line])
                    
                    if page_num in page_tables:
                        sheet_data.append([''])
                        sheet_data.append(['=== TABLES ==='])
                        sheet_data.append([''])
                
                # 페이지 테이블들 추가
                if page_num in page_tables:
                    tables_in_page = page_tables[page_num]
                    
                    for idx, table in enumerate(tables_in_page):
                        df = table.df
                        df = clean_dataframe(df)
                        
                        if len(tables_in_page) > 1:
                            sheet_data.append([clean_text_for_excel(f'Table {idx + 1}')])
                            sheet_data.append([''])
                        
                        if not df.empty:
                            for _, row in df.iterrows():
                                cleaned_row = [clean_text_for_excel(str(cell)) for cell in row.tolist()]
                                sheet_data.append(cleaned_row)
                        
                        if idx < len(tables_in_page) - 1:
                            sheet_data.append([''])
                
                # 시트에 데이터 저장
                if not sheet_data:
                    sheet_data = [['페이지에서 추출할 수 있는 내용이 없습니다.']]
                
                max_cols = max(len(row) for row in sheet_data) if sheet_data else 1
                
                for row in sheet_data:
                    while len(row) < max_cols:
                        row.append('')
                
                final_df = pd.DataFrame(sheet_data)
                sheet_name = clean_text_for_excel(f"Page_{page_num}")[:31]
                
                final_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        
        print_progress("✓ Excel 파일 생성 완료")
        return True
        
    except Exception as e:
        print_progress(f"✗ Excel 생성 실패: {str(e)}")
        return False

def create_bilingual_excel(input_xlsx, output_xlsx):
    """기존 Excel 파일에 번역 결과를 추가하여 이중 언어 Excel 생성"""
    try:
        print_progress("🌐 번역 및 이중언어 Excel 생성 중...")
        original_workbook = load_workbook(input_xlsx)
        
        with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
            
            for sheet_name in original_workbook.sheetnames:
                sheet = original_workbook[sheet_name]
                print_progress(f"처리 중인 시트: {sheet_name}")
                
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(list(row))
                
                if data:
                    bilingual_data = []
                    cell_count = 0
                    translated_count = 0
                    
                    for row in data:
                        bilingual_row = []
                        for cell in row:
                            cell_count += 1
                            if cell and isinstance(cell, str):
                                if re.search(r'[A-Za-z]', cell):
                                    translated = translate_with_deepl(cell, "EN", "KO")
                                    bilingual_row.append(f"{cell} ({translated})")
                                    translated_count += 1
                                    time.sleep(0.1)  # API 제한
                                    
                                    # 진행 상황 출력 (10개마다)
                                    if translated_count % 10 == 0:
                                        print_progress(f"번역 진행: {translated_count}개 셀 완료")
                                else:
                                    bilingual_row.append(cell)
                            else:
                                bilingual_row.append(cell)
                        bilingual_data.append(bilingual_row)
                    
                    bilingual_df = pd.DataFrame(bilingual_data)
                    sheet_name_clean = clean_text_for_excel(f"{sheet_name}_Bilingual")[:31]
                    bilingual_df.to_excel(writer, sheet_name=sheet_name_clean, 
                                         index=False, header=False)
                    
                    print_progress(f"시트 '{sheet_name}': {translated_count}/{cell_count} 셀 번역됨")
        
        print_progress("✓ 이중언어 Excel 생성 완료")
        return True
        
    except Exception as e:
        print_progress(f"✗ 번역 실패: {str(e)}")
        return False

def cleanup_temp_files(temp_files):
    """임시 파일들 정리"""
    for temp_file in temp_files:
        try:
            if os.path.exists(temp_file):
                os.unlink(temp_file)
                print_progress(f"임시 파일 삭제: {os.path.basename(temp_file)}")
        except Exception as e:
            print_progress(f"임시 파일 삭제 실패: {os.path.basename(temp_file)} - {str(e)}")
            pass

def process_single_pdf(input_pdf, output_directory):
    """단일 PDF 파일 처리"""
    filename = os.path.basename(input_pdf)
    base_name = os.path.splitext(filename)[0]
    temp_files = []
    
    try:
        print_progress(f"처리 시작: {filename}", is_header=True)
        
        # 1단계: PDF 내용 사전 확인
        print_progress("1단계: PDF 내용 분석")
        has_tables, has_text, text_length, content_score = check_pdf_has_tables_and_text(input_pdf)
        
        # 2단계: 조건부 OCR 처리
        print_progress("2단계: OCR 필요성 판단")
        pdf_to_process = input_pdf
        
        skip_ocr = False
        if has_tables:
            skip_ocr = True
            reason = "유효한 테이블이 발견됨"
        elif has_text and text_length > 500:
            skip_ocr = True
            reason = "충분한 텍스트 내용이 발견됨"
        elif content_score >= 70:
            skip_ocr = True
            reason = "내용 품질이 충분함"
        
        if skip_ocr:
            print_progress(f"📄 OCR 건너뛰기: {reason}")
        else:
            print_progress(f"🔍 OCR 처리 필요 (점수: {content_score}/150)")
            ocr_pdf = ocr_pdf_with_options(input_pdf)
            if ocr_pdf != input_pdf:
                pdf_to_process = ocr_pdf
                temp_files.append(ocr_pdf)
        
        # 3단계: Excel 추출
        print_progress("3단계: Excel 추출")
        temp_excel = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
        temp_files.append(temp_excel)
        
        success = pdf_to_excel(pdf_to_process, temp_excel)
        if not success:
            print_progress("✗ Excel 추출 실패")
            return False
        
        # 4단계: 번역 및 최종 파일 생성
        print_progress("4단계: 번역 및 최종 파일 생성")
        final_output = get_unique_filename(output_directory, base_name)
        success = create_bilingual_excel(temp_excel, final_output)
        
        if success:
            print_progress(f"✅ 처리 완료: {os.path.basename(final_output)}")
        else:
            print_progress("✗ 번역 실패")
        
        # 5단계: 임시 파일 정리
        print_progress("5단계: 임시 파일 정리")
        cleanup_temp_files(temp_files)
        
        return success
        
    except Exception as e:
        print_progress(f"✗ 처리 중 오류: {str(e)}")
        cleanup_temp_files(temp_files)
        return False

def main():
    """메인 실행 함수 - OCR 의존성 확인 추가"""
    current_directory = os.getcwd()
    
    print_progress("PDF 번역기 시작", is_header=True)
    print_progress(f"작업 디렉토리: {current_directory}")
    
    # OCR 환경 사전 확인
    print_progress("OCR 환경 확인", is_header=True)
    ocr_available = check_ocr_dependencies()
    if ocr_available:
        available_lang = check_tesseract_languages()
        print_progress(f"사용할 OCR 언어: {available_lang}")
    else:
        print_progress("⚠ OCR이 사용 불가능합니다. 텍스트 추출만 진행됩니다.")
    
    # 현재 디렉토리에서 PDF 파일 찾기
    pdf_files = find_pdf_files(current_directory)
    
    if not pdf_files:
        print_progress("❌ PDF 파일을 찾을 수 없습니다.")
        input("Enter 키를 눌러 종료하세요...")
        return
    
    print_progress(f"📁 발견된 PDF 파일: {len(pdf_files)}개")
    for i, pdf_file in enumerate(pdf_files, 1):
        print_progress(f"  {i}. {os.path.basename(pdf_file)}")
    
    # 처리 시작
    success_count = 0
    total_files = len(pdf_files)
    
    for i, pdf_file in enumerate(pdf_files, 1):
        print_progress(f"\n진행률: {i}/{total_files}", is_header=True)
        
        success = process_single_pdf(pdf_file, current_directory)
        if success:
            success_count += 1
    
    # 최종 결과
    print_progress("작업 완료", is_header=True)
    print_progress(f"🎉 전체 결과: {success_count}/{total_files} 성공")
    
    if success_count > 0:
        print_progress("✅ 생성된 파일들:")
        for filename in os.listdir(current_directory):
            if filename.endswith('_final.xlsx'):
                print_progress(f"  📄 {filename}")
    
    print_progress("프로그램을 종료합니다.")
    input("Enter 키를 눌러 종료하세요...")

if __name__ == '__main__':
    main()