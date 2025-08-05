#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PDF 번역기 - 메인 실행 파일
PDF 파일에서 텍스트와 테이블을 추출하고 DeepL API로 번역하여 Excel 파일로 저장합니다.
"""

import camelot
import pdfplumber
import pandas as pd
import re
import html
import requests
import time
import datetime
import subprocess
import os
import tempfile
import sys
from openpyxl import load_workbook

# 콘솔 인코딩 설정 (Windows)
if sys.platform.startswith('win'):
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# DeepL API 설정
DEEPL_API_KEY = "b3125acc-3a44-4648-8b4d-5ca8e7350059:fx"
DEEPL_API_URL = "https://api-free.deepl.com/v2/translate"

def print_progress(message, is_header=False):
    """진행 상황 출력"""
    timestamp = datetime.datetime.now().strftime("%H:%M:%S")
    if is_header:
        print("=" * 60)
        print(f"[{timestamp}] {message}")
        print("=" * 60)
    else:
        print(f"[{timestamp}] {message}")

def get_unique_filename(directory, base_name, extension=".xlsx"):
    """고유한 파일명 생성"""
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{base_name}_{timestamp}_번역완료{extension}"
    return os.path.join(directory, filename)

def find_pdf_files(directory):
    """디렉토리에서 모든 PDF 파일 찾기"""
    pdf_files = []
    for filename in os.listdir(directory):
        if filename.lower().endswith('.pdf'):
            pdf_files.append(os.path.join(directory, filename))
    return sorted(pdf_files)

def check_pdf_has_tables_and_text(input_pdf):
    """PDF에 테이블과 텍스트가 있는지 사전 확인"""
    has_tables = False
    has_text = False
    total_text_length = 0
    
    try:
        # 테이블 확인
        tables = camelot.read_pdf(input_pdf, pages='all')
        if tables and len(tables) > 0:
            valid_tables = 0
            for table in tables:
                if table.df.shape[0] > 1 and table.df.shape[1] > 1:
                    non_empty_cells = table.df.count().sum()
                    total_cells = table.df.shape[0] * table.df.shape[1]
                    if non_empty_cells / total_cells > 0.1:
                        valid_tables += 1
            
            if valid_tables > 0:
                has_tables = True
                print_progress(f"✓ 유효한 테이블 {valid_tables}개 발견")
    except Exception:
        # Camelot 실패 시 pdfplumber로 테이블 감지 시도
        try:
            with pdfplumber.open(input_pdf) as pdf:
                for page in pdf.pages:
                    page_tables = page.find_tables()
                    if page_tables and len(page_tables) > 0:
                        has_tables = True
                        print_progress("✓ 테이블 발견")
                        break
        except:
            pass
    
    # 텍스트 확인
    try:
        with pdfplumber.open(input_pdf) as pdf:
            for page in pdf.pages:
                text_methods = [
                    page.extract_text(),
                    page.extract_text(layout=True),
                    page.extract_text(x_tolerance=3, y_tolerance=3)
                ]
                
                for text in text_methods:
                    if text:
                        clean_text = text.strip()
                        unique_chars = len(set(clean_text.replace(' ', '').replace('\n', '')))
                        if len(clean_text) > 50 and unique_chars > 10:
                            total_text_length += len(clean_text)
                            has_text = True
        
        if has_text:
            print_progress(f"✓ 텍스트 발견 (총 {total_text_length} 문자)")
        else:
            print_progress("✗ 의미있는 텍스트 없음")
            
    except Exception:
        print_progress("✗ 텍스트 추출 실패")
    
    # 종합 점수 계산
    content_score = 0
    if has_tables:
        content_score += 100
    if has_text and total_text_length > 100:
        content_score += 50
    elif has_text:
        content_score += 20
    
    return has_tables, has_text, total_text_length, content_score

def translate_with_deepl(text, source_lang="EN", target_lang="KO"):
    """DeepL API를 사용하여 텍스트 번역"""
    if not text or not text.strip():
        return ""
    
    # 너무 긴 텍스트는 잘라서 번역
    if len(text) > 1000:
        text = text[:1000] + "..."
    
    headers = {
        "Authorization": f"DeepL-Auth-Key {DEEPL_API_KEY}",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    
    data = {
        "text": text,
        "source_lang": source_lang,
        "target_lang": target_lang
    }
    
    try:
        response = requests.post(DEEPL_API_URL, headers=headers, data=data, timeout=10)
        
        if response.status_code == 200:
            result = response.json()
            return result["translations"][0]["text"]
        else:
            print_progress(f"번역 API 오류: {response.status_code}")
            return text
    except Exception as e:
        print_progress(f"번역 중 오류: {str(e)}")
        return text

def clean_text_for_excel(text):
    """Excel XML 호환을 위한 텍스트 정리"""
    if not text:
        return ""
    
    text = str(text)
    text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', text)
    text = html.unescape(text)
    text = text.replace('&', '&amp;')
    text = text.replace('<', '&lt;')
    text = text.replace('>', '&gt;')
    text = text.replace('"', '&quot;')
    text = text.replace("'", '&apos;')
    text = re.sub(r'\s+', ' ', text)
    text = text.strip()
    
    return text

def extract_text_improved(text):
    """텍스트 추출 및 정리 함수 개선"""
    if not text:
        return ""
    
    text = text.strip()
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'\n\s*\n', '\n', text)
    cleaned_text = re.sub(r'[^\w\s.,!?;:\-\(\)\[\]\/\\\'"@#$%&+=<>]', ' ', text)
    cleaned_text = re.sub(r'\s+', ' ', cleaned_text)
    cleaned_text = clean_text_for_excel(cleaned_text)
    
    return cleaned_text.strip()

def clean_dataframe(df):
    """DataFrame의 모든 셀을 Excel 호환 형태로 정리"""
    if df.empty:
        return df
    
    cleaned_df = df.copy()
    for col in cleaned_df.columns:
        cleaned_df[col] = cleaned_df[col].apply(
            lambda x: clean_text_for_excel(str(x)) if pd.notna(x) else ""
        )
    
    return cleaned_df

def pdf_to_excel(input_pdf, output_xlsx):
    """PDF에서 텍스트와 테이블을 추출하여 Excel로 변환"""
    try:
        print_progress("📊 테이블 추출 중...")
        # 테이블 추출
        tables = []
        try:
            tables = camelot.read_pdf(input_pdf, pages='all')
            if tables:
                print_progress(f"✓ {len(tables)}개 테이블 추출됨")
        except Exception:
            print_progress("✗ 테이블 추출 실패")
        
        print_progress("📄 텍스트 추출 중...")
        # 텍스트 추출
        page_texts = {}
        total_text_length = 0
        
        with pdfplumber.open(input_pdf) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                text_methods = [
                    page.extract_text(),
                    page.extract_text(layout=True),
                    page.extract_text(x_tolerance=3, y_tolerance=3),
                    page.extract_text(y_tolerance=3)
                ]
                
                best_text = ""
                for text in text_methods:
                    if text and len(text) > len(best_text):
                        best_text = text
                
                if best_text:
                    cleaned_text = extract_text_improved(best_text)
                    page_texts[page_num] = cleaned_text
                    total_text_length += len(cleaned_text)
        
        print_progress(f"✓ {len(page_texts)}개 페이지에서 {total_text_length} 문자 추출됨")
        
        # 페이지별로 테이블 그룹화
        page_tables = {}
        if tables:
            for table in tables:
                page_num = table.page
                if page_num not in page_tables:
                    page_tables[page_num] = []
                page_tables[page_num].append(table)
        
        print_progress("📋 Excel 파일 생성 중...")
        # Excel 파일 생성
        with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
            all_pages = set(page_tables.keys()) | set(page_texts.keys())
            
            if not all_pages:
                all_pages = {1}
                page_texts[1] = "PDF에서 내용을 추출할 수 없습니다."
            
            for page_num in sorted(all_pages):
                sheet_data = []
                
                # 페이지 텍스트 추가
                if page_num in page_texts and page_texts[page_num]:
                    text_content = page_texts[page_num]
                    text_lines = text_content.split('\n')
                    
                    for line in text_lines:
                        line = line.strip()
                        if line:
                            cleaned_line = clean_text_for_excel(line)
                            if cleaned_line:
                                sheet_data.append([cleaned_line])
                    
                    if page_num in page_tables:
                        sheet_data.append([''])
                        sheet_data.append(['=== TABLES ==='])
                        sheet_data.append([''])
                
                # 페이지 테이블들 추가
                if page_num in page_tables:
                    tables_in_page = page_tables[page_num]
                    
                    for idx, table in enumerate(tables_in_page):
                        df = table.df
                        df = clean_dataframe(df)
                        
                        if len(tables_in_page) > 1:
                            sheet_data.append([clean_text_for_excel(f'Table {idx + 1}')])
                            sheet_data.append([''])
                        
                        if not df.empty:
                            for _, row in df.iterrows():
                                cleaned_row = [clean_text_for_excel(str(cell)) for cell in row.tolist()]
                                sheet_data.append(cleaned_row)
                        
                        if idx < len(tables_in_page) - 1:
                            sheet_data.append([''])
                
                # 시트에 데이터 저장
                if not sheet_data:
                    sheet_data = [['페이지에서 추출할 수 있는 내용이 없습니다.']]
                
                max_cols = max(len(row) for row in sheet_data) if sheet_data else 1
                
                for row in sheet_data:
                    while len(row) < max_cols:
                        row.append('')
                
                final_df = pd.DataFrame(sheet_data)
                sheet_name = clean_text_for_excel(f"Page_{page_num}")[:31]
                
                final_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        
        print_progress("✓ Excel 파일 생성 완료")
        return True
        
    except Exception as e:
        print_progress(f"✗ Excel 생성 실패: {str(e)}")
        return False

def create_bilingual_excel(input_xlsx, output_xlsx):
    """기존 Excel 파일에 번역 결과를 추가하여 이중 언어 Excel 생성"""
    try:
        print_progress("🌐 번역 및 이중언어 Excel 생성 중...")
        original_workbook = load_workbook(input_xlsx)
        
        with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
            
            for sheet_name in original_workbook.sheetnames:
                sheet = original_workbook[sheet_name]
                
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(list(row))
                
                if data:
                    bilingual_data = []
                    cell_count = 0
                    translated_count = 0
                    
                    for row in data:
                        bilingual_row = []
                        for cell in row:
                            cell_count += 1
                            if cell and isinstance(cell, str) and re.search(r'[A-Za-z]', cell):
                                # 영어 텍스트가 있으면 번역 추가
                                translated = translate_with_deepl(cell, "EN", "KO")
                                bilingual_row.append(f"{cell}\n({translated})")
                                translated_count += 1
                                time.sleep(0.1)  # API 제한
                                
                                # 진행 상황 출력 (5개마다)
                                if translated_count % 5 == 0:
                                    print_progress(f"번역 진행: {translated_count}개 셀 완료")
                            else:
                                bilingual_row.append(cell if cell else "")
                        bilingual_data.append(bilingual_row)
                    
                    bilingual_df = pd.DataFrame(bilingual_data)
                    sheet_name_clean = clean_text_for_excel(f"{sheet_name}_번역")[:31]
                    bilingual_df.to_excel(writer, sheet_name=sheet_name_clean, 
                                         index=False, header=False)
                    
                    print_progress(f"시트 '{sheet_name}': {translated_count}/{cell_count} 셀 번역됨")
        
        print_progress("✓ 이중언어 Excel 생성 완료")
        return True
        
    except Exception as e:
        print_progress(f"✗ 번역 실패: {str(e)}")
        return False

def cleanup_temp_files(temp_files):
    """임시 파일들 정리"""
    for temp_file in temp_files:
        try:
            if os.path.exists(temp_file):
                os.unlink(temp_file)
                print_progress(f"임시 파일 삭제: {os.path.basename(temp_file)}")
        except:
            pass

def process_single_pdf(input_pdf, output_directory):
    """단일 PDF 파일 처리"""
    filename = os.path.basename(input_pdf)
    base_name = os.path.splitext(filename)[0]
    temp_files = []
    
    try:
        print_progress(f"처리 시작: {filename}", is_header=True)
        
        # 1단계: PDF 내용 사전 확인
        print_progress("1단계: PDF 내용 분석")
        has_tables, has_text, text_length, content_score = check_pdf_has_tables_and_text(input_pdf)
        
        # 2단계: Excel 추출 (OCR 생략)
        print_progress("2단계: Excel 추출")
        temp_excel = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
        temp_files.append(temp_excel)
        
        success = pdf_to_excel(input_pdf, temp_excel)
        if not success:
            print_progress("✗ Excel 추출 실패")
            return False
        
        # 3단계: 번역 및 최종 파일 생성
        print_progress("3단계: 번역 및 최종 파일 생성")
        final_output = get_unique_filename(output_directory, base_name)
        success = create_bilingual_excel(temp_excel, final_output)
        
        if success:
            print_progress(f"✅ 처리 완료: {os.path.basename(final_output)}")
            print_progress(f"📁 저장 위치: {final_output}")
        else:
            print_progress("✗ 번역 실패")
        
        # 4단계: 임시 파일 정리
        cleanup_temp_files(temp_files)
        
        return success
        
    except Exception as e:
        print_progress(f"✗ 처리 중 오류: {str(e)}")
        cleanup_temp_files(temp_files)
        return False

def main():
    """메인 실행 함수"""
    current_directory = os.getcwd()
    
    print_progress("🔄 PDF 번역기 시작", is_header=True)
    print_progress(f"📂 작업 디렉토리: {current_directory}")
    print_progress("📝 이 프로그램은 PDF 파일을 읽어서 영어를 한국어로 번역하여 Excel 파일로 저장합니다.")
    
    # 현재 디렉토리에서 PDF 파일 찾기
    pdf_files = find_pdf_files(current_directory)
    
    if not pdf_files:
        print_progress("❌ PDF 파일을 찾을 수 없습니다.")
        print_progress("💡 이 프로그램과 같은 폴더에 PDF 파일을 넣어주세요.")
        input("\n아무 키나 누르면 종료됩니다...")
        return
    
    print_progress(f"📁 발견된 PDF 파일: {len(pdf_files)}개")
    for i, pdf_file in enumerate(pdf_files, 1):
        print_progress(f"  {i}. {os.path.basename(pdf_file)}")
    
    # 처리 시작
    success_count = 0
    total_files = len(pdf_files)
    
    for i, pdf_file in enumerate(pdf_files, 1):
        print_progress(f"\n📄 진행률: {i}/{total_files}", is_header=True)
        
        success = process_single_pdf(pdf_file, current_directory)
        if success:
            success_count += 1
    
    # 최종 결과
    print_progress("🎉 작업 완료!", is_header=True)
    print_progress(f"✅ 전체 결과: {success_count}/{total_files} 성공")
    
    if success_count > 0:
        print_progress("📄 생성된 파일들:")
        for filename in os.listdir(current_directory):
            if filename.endswith('_번역완료.xlsx'):
                print_progress(f"  📋 {filename}")
    
    print_progress("🔚 프로그램을 종료합니다.")
    input("\n아무 키나 누르면 창이 닫힙니다...")

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"프로그램 실행 중 오류가 발생했습니다: {str(e)}")
        input("\n아무 키나 누르면 창이 닫힙니다...")
